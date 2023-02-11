import os
import re
import shutil
from random import randint, choice
import random

import openpyxl
from openpyxl.utils import column_index_from_string

def co_po_map_optimize(f_path, wb, ws, start_col, end_col, start_row, end_row):
    ind = 0
    n_count = 0
    for col in range(start_col, end_col + 1):
        ind += 1
        col_letter = openpyxl.utils.cell.get_column_letter(col)
        for row in range(start_row, end_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value == 1 or 0:
                n_count += 1
                # print(cell.value)
                # cell.value = randint('', 2)
                cell.value = ''
    print(f"Optimizing Co=PO Mapping by replacing 0 and 1 with Blank | Total {n_count} replaces")
    wb.save(f_path)

def no_of_co_correction(f_path, wb, ws, start_col, end_col, start_row, end_row):
    ind = 0
    co_count = 0
    print(str(start_col) + " | " + str(end_col))
    for row in range(start_row, end_row + 1):
        ind += 1
        non_zero_count = 0
        for col in range(start_col, end_col + 1):
            col_letter = openpyxl.utils.cell.get_column_letter(col)
            cell = ws[f"{col_letter}{row}"]
            if cell.value != '' or 0 or None:
                # print(cell.value)
                if cell.value != None:
                    non_zero_count = non_zero_count + int(cell.value)
                # print(cell.value)
            # print("-----------------")
        if non_zero_count > 0:
            co_count += 1
    print(f"Total no of COs = {co_count}")
    return co_count
    # wb.save(f_path)

def no_of_co_replace(f_path, wb, ws, col, row, no_of_co):
    ind = 0
    col_letter = openpyxl.utils.cell.get_column_letter(col)
    cell = ws[f"{col_letter}{row}"]
    if cell.value != no_of_co:
        print(f"No of COs in GENERAL INPUT Sheet corrected from {cell.value} to {no_of_co} !!!")
    else:
        print(f"No of COs in GENERAL INPUT Sheet Replaced to {no_of_co}")
    cell.value = no_of_co
    wb.save(f_path)

def find_non_zero_count(f_path, wb, ws, start_col, end_col, start_row, end_row, result_row):
    ind = 0
    for col in range(start_col, end_col + 1):
        ind += 1
        col_letter = openpyxl.utils.cell.get_column_letter(col)
        non_zero_count = 0
        for row in range(start_row, end_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value != 0 or None or '':
                non_zero_count += 1
        result_cell = ws[f"{col_letter}{result_row}"]
        result_cell.value = f"=COUNTIF({col_letter}{start_row}:{col_letter}{end_row},\"<>0\")"
        print(f"Counting non zero COs for PO-{ind}")
    wb.save(f_path)

def change_cell_values(f_path, wb, sheet_n, items, start_row, end_row, column):
    for i, item in enumerate(items):
        row = start_row + i
        if row > end_row:
            break
        sheet_n.cell(row=row, column=column, value=item)
        print(f"Updating Equation PO-{i+1}")
    wb.save(f_path)

def change_direct_attainment_internal(f_path, wb, ws, start_col, end_col, start_row, end_row, value_list):
    ind = 0
    print("Changing Direct Attainment - Internal-----!!!")
    for col in range(start_col, end_col + 1):
        col_letter = openpyxl.utils.cell.get_column_letter(col)
        non_zero_count = 0
        for row in range(start_row, end_row + 1):
            cell = ws[f"{col_letter}{row}"]
            cell.value = value_list[ind]
            ind += 1
    wb.save(f_path)

def change_direct_attainment_university(f_path, wb, ws, start_col, end_col, start_row, end_row, value_list):
    ind = 0
    print("Changing Direct Attainment - University-----!!!")
    for col in range(start_col, end_col + 1):
        col_letter = openpyxl.utils.cell.get_column_letter(col)
        non_zero_count = 0
        for row in range(start_row, end_row + 1):
            cell = ws[f"{col_letter}{row}"]
            cell.value = value_list[ind]
            ind += 1
    wb.save(f_path)

def extract_class_and_sub_code_from_filename(filename):
    print(filename)
    # class_name = re.compile(r'{S}\d[3-8]{ CS }\d{1,2}')
    class_name = re.compile(r'[S]\d{1}[_]*\D{2,3}\d{1}')
    class_n = class_name.search(filename)
    class_nam = class_n.group().replace("_",'')

    sub_code = re.compile(r'[CS]{2}\d{3}')
    sub_c = sub_code.search(filename)
    sub_cod = sub_c.group().replace("_",'')

    print(f"Class Name : {class_nam}")
    print(f"Subject Code : {sub_cod}")
    return [class_nam, sub_cod]


import openpyxl

def copy_cells_PO_PSO_Actual(wb, dest_file, src_ws, dest_sheet_prefix, extra_col1, extra_col2, cells_to_copy):
    # Load the source workbook and select the source sheet

    # openpyxl.Workbook.close(wb)
    # src_wb = openpyxl.load_workbook(wb, data_only=True)
    #
    # # src_wb = wb
    src_sheet_PO = src_ws

    # Load the destination workbook
    dest_wb = openpyxl.load_workbook(dest_file)

    # Get the next available sheet name with the given prefix
    dest_sheet = dest_wb[dest_sheet_prefix]

    # Find the next available row in the destination sheet
    dest_row = dest_sheet.max_row + 1
    print(dest_row)
    # Copy the values from source to destination, including the extra column values
    dest_sheet.cell(row=dest_row, column=1).value = extra_col1
    dest_sheet.cell(row=dest_row, column=2).value = extra_col2
    dest_sheet.cell(row=dest_row, column=3).value = src_sheet_PO['D22'].value
    print(src_sheet_PO['D22'].value)
    dest_sheet.cell(row=dest_row, column=4).value = src_sheet_PO['D23'].value
    print(src_sheet_PO['D23'].value)
    dest_sheet.cell(row=dest_row, column=5).value = src_sheet_PO['D23'].value
    print(src_sheet_PO['D24'].value)
    dest_sheet.cell(row=dest_row, column=6).value = src_sheet_PO['D24'].value
    print(src_sheet_PO['D25'].value)
    dest_sheet.cell(row=dest_row, column=7).value = src_sheet_PO['D25'].value
    print(src_sheet_PO['D26'].value)
    dest_sheet.cell(row=dest_row, column=8).value = src_sheet_PO['D26'].value
    print(src_sheet_PO['D27'].value)
    dest_sheet.cell(row=dest_row, column=9).value = src_sheet_PO['D27'].value
    print(src_sheet_PO['D28'].value)
    dest_sheet.cell(row=dest_row, column=10).value = src_sheet_PO['D28'].value
    print(src_sheet_PO['D29'].value)
    dest_sheet.cell(row=dest_row, column=11).value = src_sheet_PO['D29'].value
    print(src_sheet_PO['D30'].value)
    dest_sheet.cell(row=dest_row, column=12).value = src_sheet_PO['D30'].value
    print(src_sheet_PO['D31'].value)
    dest_sheet.cell(row=dest_row, column=13).value = src_sheet_PO['D31'].value
    print(src_sheet_PO['D32'].value)
    dest_sheet.cell(row=dest_row, column=14).value = src_sheet_PO['D32'].value
    print(src_sheet_PO['D32'].value)
    dest_sheet.cell(row=dest_row, column=15).value = src_sheet_PO['D33'].value
    print(src_sheet_PO['D33'].value)
    # dest_col = 3
    # for coord in cells_to_copy:
    #     dest_col = openpyxl.utils.column_index_from_string(coord[0]) + 2
    #     dest_sheet.cell(row=dest_row, column=dest_col).value = src_sheet[cells_to_copy[index]].value
    #     dest_col += 1
    #     print(dest_col)
    #     print(src_sheet[cells_to_copy[index]].value)
    # Save the destination workbook
    dest_wb.save(dest_file)


# Example usage:
folder = os.getcwd()
# folder_p = folder + "\\" + "Files_Under_Processing"
folder_p = folder
folder_path = folder_p  # change to the actual path of your folder
password = 'NithinNBACSE'  # change to the password to unprotect the sheet
sheet_name = 'PO Attainment'  # change to the name of the sheet to change
sheet_name_co_po = 'PO Input'  # change to the name of the sheet to change
sheet_name_gen_in = 'GENERAL INPUT'  # change to the name of the sheet to change
direct_attain_internal = [60, 60, 70, 80]
direct_attain_university = [60, 70, 80]
formulas = ["=IFERROR(ROUND((((D$5*'TOTAL CO Attainment'!$C$5)+(D$6*'TOTAL CO Attainment'!$C$6)+(D$7*'TOTAL CO Attainment'!$C$7)+(D$8*'TOTAL CO Attainment'!$C$8)+(D$9*'TOTAL CO Attainment'!$C$9)+(D$10*'TOTAL CO Attainment'!$C$10)+(D$11*'TOTAL CO Attainment'!$C$11)+(D$12*'TOTAL CO Attainment'!$C$12)+(D$13*'TOTAL CO Attainment'!$C$13)+(D$14*'TOTAL CO Attainment'!$C$14))/(3*D18)),2),"")",

"=IFERROR(ROUND((((E$5*'TOTAL CO Attainment'!$C$5)+(E$6*'TOTAL CO Attainment'!$C$6)+(E$7*'TOTAL CO Attainment'!$C$7)+(E$8*'TOTAL CO Attainment'!$C$8)+(E$9*'TOTAL CO Attainment'!$C$9)+(E$10*'TOTAL CO Attainment'!$C$10)+(E$11*'TOTAL CO Attainment'!$C$11)+(E$12*'TOTAL CO Attainment'!$C$12)+(E$13*'TOTAL CO Attainment'!$C$13)+(E$14*'TOTAL CO Attainment'!$C$14))/(3*E18)),2),"")",

"=IFERROR(ROUND((((F$5*'TOTAL CO Attainment'!$C$5)+(F$6*'TOTAL CO Attainment'!$C$6)+(F$7*'TOTAL CO Attainment'!$C$7)+(F$8*'TOTAL CO Attainment'!$C$8)+(F$9*'TOTAL CO Attainment'!$C$9)+(F$10*'TOTAL CO Attainment'!$C$10)+(F$11*'TOTAL CO Attainment'!$C$11)+(F$12*'TOTAL CO Attainment'!$C$12)+(F$13*'TOTAL CO Attainment'!$C$13)+(F$14*'TOTAL CO Attainment'!$C$14))/(3*F18)),2),"")",

"=IFERROR(ROUND((((G$5*'TOTAL CO Attainment'!$C$5)+(G$6*'TOTAL CO Attainment'!$C$6)+(G$7*'TOTAL CO Attainment'!$C$7)+(G$8*'TOTAL CO Attainment'!$C$8)+(G$9*'TOTAL CO Attainment'!$C$9)+(G$10*'TOTAL CO Attainment'!$C$10)+(G$11*'TOTAL CO Attainment'!$C$11)+(G$12*'TOTAL CO Attainment'!$C$12)+(G$13*'TOTAL CO Attainment'!$C$13)+(G$14*'TOTAL CO Attainment'!$C$14))/(3*G18)),2),"")",

"=IFERROR(ROUND((((H$5*'TOTAL CO Attainment'!$C$5)+(H$6*'TOTAL CO Attainment'!$C$6)+(H$7*'TOTAL CO Attainment'!$C$7)+(H$8*'TOTAL CO Attainment'!$C$8)+(H$9*'TOTAL CO Attainment'!$C$9)+(H$10*'TOTAL CO Attainment'!$C$10)+(H$11*'TOTAL CO Attainment'!$C$11)+(H$12*'TOTAL CO Attainment'!$C$12)+(H$13*'TOTAL CO Attainment'!$C$13)+(H$14*'TOTAL CO Attainment'!$C$14))/(3*H18)),2),"")",

"=IFERROR(ROUND((((I$5*'TOTAL CO Attainment'!$C$5)+(I$6*'TOTAL CO Attainment'!$C$6)+(I$7*'TOTAL CO Attainment'!$C$7)+(I$8*'TOTAL CO Attainment'!$C$8)+(I$9*'TOTAL CO Attainment'!$C$9)+(I$10*'TOTAL CO Attainment'!$C$10)+(I$11*'TOTAL CO Attainment'!$C$11)+(I$12*'TOTAL CO Attainment'!$C$12)+(I$13*'TOTAL CO Attainment'!$C$13)+(I$14*'TOTAL CO Attainment'!$C$14))/(3*I18)),2),"")",

"=IFERROR(ROUND((((J$5*'TOTAL CO Attainment'!$C$5)+(J$6*'TOTAL CO Attainment'!$C$6)+(J$7*'TOTAL CO Attainment'!$C$7)+(J$8*'TOTAL CO Attainment'!$C$8)+(J$9*'TOTAL CO Attainment'!$C$9)+(J$10*'TOTAL CO Attainment'!$C$10)+(J$11*'TOTAL CO Attainment'!$C$11)+(J$12*'TOTAL CO Attainment'!$C$12)+(J$13*'TOTAL CO Attainment'!$C$13)+(J$14*'TOTAL CO Attainment'!$C$14))/(3*J18)),2),"")",

"=IFERROR(ROUND((((K$5*'TOTAL CO Attainment'!$C$5)+(K$6*'TOTAL CO Attainment'!$C$6)+(K$7*'TOTAL CO Attainment'!$C$7)+(K$8*'TOTAL CO Attainment'!$C$8)+(K$9*'TOTAL CO Attainment'!$C$9)+(K$10*'TOTAL CO Attainment'!$C$10)+(K$11*'TOTAL CO Attainment'!$C$11)+(K$12*'TOTAL CO Attainment'!$C$12)+(K$13*'TOTAL CO Attainment'!$C$13)+(K$14*'TOTAL CO Attainment'!$C$14))/(3*K18)),2),"")",

"=IFERROR(ROUND((((L$5*'TOTAL CO Attainment'!$C$5)+(L$6*'TOTAL CO Attainment'!$C$6)+(L$7*'TOTAL CO Attainment'!$C$7)+(L$8*'TOTAL CO Attainment'!$C$8)+(L$9*'TOTAL CO Attainment'!$C$9)+(L$10*'TOTAL CO Attainment'!$C$10)+(L$11*'TOTAL CO Attainment'!$C$11)+(L$12*'TOTAL CO Attainment'!$C$12)+(L$13*'TOTAL CO Attainment'!$C$13)+(L$14*'TOTAL CO Attainment'!$C$14))/(3*L18)),2),"")",

"=IFERROR(ROUND((((M$5*'TOTAL CO Attainment'!$C$5)+(M$6*'TOTAL CO Attainment'!$C$6)+(M$7*'TOTAL CO Attainment'!$C$7)+(M$8*'TOTAL CO Attainment'!$C$8)+(M$9*'TOTAL CO Attainment'!$C$9)+(M$10*'TOTAL CO Attainment'!$C$10)+(M$11*'TOTAL CO Attainment'!$C$11)+(M$12*'TOTAL CO Attainment'!$C$12)+(M$13*'TOTAL CO Attainment'!$C$13)+(M$14*'TOTAL CO Attainment'!$C$14))/(3*M18)),2),"")",

"=IFERROR(ROUND((((N$5*'TOTAL CO Attainment'!$C$5)+(N$6*'TOTAL CO Attainment'!$C$6)+(N$7*'TOTAL CO Attainment'!$C$7)+(N$8*'TOTAL CO Attainment'!$C$8)+(N$9*'TOTAL CO Attainment'!$C$9)+(N$10*'TOTAL CO Attainment'!$C$10)+(N$11*'TOTAL CO Attainment'!$C$11)+(N$12*'TOTAL CO Attainment'!$C$12)+(N$13*'TOTAL CO Attainment'!$C$13)+(N$14*'TOTAL CO Attainment'!$C$14))/(3*N18)),2),"")",

"=IFERROR(ROUND((((O$5*'TOTAL CO Attainment'!$C$5)+(O$6*'TOTAL CO Attainment'!$C$6)+(O$7*'TOTAL CO Attainment'!$C$7)+(O$8*'TOTAL CO Attainment'!$C$8)+(O$9*'TOTAL CO Attainment'!$C$9)+(O$10*'TOTAL CO Attainment'!$C$10)+(O$11*'TOTAL CO Attainment'!$C$11)+(O$12*'TOTAL CO Attainment'!$C$12)+(O$13*'TOTAL CO Attainment'!$C$13)+(O$14*'TOTAL CO Attainment'!$C$14))/(3*O18)),2),"")"]

if __name__ == '__main__':
    subfolder = "Files_with_Inconsistency"
    # cwd = os.getcwd()
    subfolder_path = os.path.join(folder_p, subfolder)
    if not os.path.exists(subfolder_path):
        os.makedirs(subfolder_path)
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            try:
                print(f"Processing :------------------------ {filename} ------------------------")

                # workbook = openpyxl.load_workbook(file_path)
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook[sheet_name]
                worksheet_co_po = workbook[sheet_name_co_po]
                worksheet_gen_in = workbook[sheet_name_gen_in]
                # unprotect the sheet if it's protected
                if worksheet.protection.sheet:
                    worksheet.protection.sheet = False
                    worksheet.protection.password = password

                find_non_zero_count(file_path, workbook, worksheet, 4, 15, 5, 14, 18)
                change_cell_values(file_path, workbook, worksheet, formulas, 22, 33, 4)
                # co_po_map_optimize(file_path, workbook, worksheet_co_po, 3, 14, 6, 15)
                change_direct_attainment_internal(file_path, workbook, worksheet_gen_in, 3, 3, 16, 19, direct_attain_internal)
                change_direct_attainment_university(file_path, workbook, worksheet_gen_in, 3, 3, 25, 27, direct_attain_university)
                no_of_co = no_of_co_correction(file_path, workbook, worksheet_co_po, 3, 14, 6, 15)
                no_of_co_replace(file_path, workbook, worksheet_gen_in, 3, 13, no_of_co)

                # [class_name, sub_code] = extract_class_and_sub_code_from_filename(filename)
                # # openpyxl.Workbook.save(file_path)
                # copy_cells_PO_PSO_Actual(file_path, 'Final_Attainment_Report.xlsx', worksheet, 'PO and PSO Actual', class_name, sub_code,
                #            ['D22', 'D23', 'D24', 'D25', 'D26', 'D27', 'D28', 'D29', 'D30', 'D31', 'D32', 'D33'])
                # If processing is successful, do nothing
                # workbook.save()
            except Exception as e:
                print(e)
                # If an error occurs, move the file to the subfolder
                shutil.move(file_path, subfolder_path)
